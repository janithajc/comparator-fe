import csv

import sys

import time

import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
import xml.etree.ElementTree as ET
from decimal import Decimal

datacompareFooter= ""
datacompareDataFormat=""

colsToConsider=["DC #","Product #"]  #put pk/compositekey to consider
#colsToConsider=["DC #","Product #"]  #put pk/compositekey to consider
#colsToConsider=["DC #" ,"Ship-To #"]  #put pk/compositekey to consider
secondaryPk=["DC #","DC #"]
Stringcol = ["MPC","GTIN"]
castToInt = ["GTIN"]

class WriteToExcel():
    workSheet=None
    columnCode=1
    rowCode=1
    wb=None
    colourCode={"Matched": "00ff00", "SlightlyMatched": "ffff00", "MissMatched": "ff0000", "Headers": "D3D3D3", "PreviouslyKnownMismatched": "F5AE14", "FoundInOneFile": "C39BD3", "FoundInTwoFile": "168FC9"}
    colType = {'168FC9': 'FoundInTwoFile', 'C39BD3': 'FoundInOneFile', '00ff00': 'Matched', 'ffff00': 'SlightlyMatched',
               'D3D3D3': 'Headers', 'ff0000': 'MissMatched', 'F5AE14': 'PreviouslyKnownMismatched'}
    colourCodeDesc={"Matched": "Matched records in both files", "SlightlyMatched": "Slightly Matched records in both files", "MissMatched": "MisMatched records in both files", "Headers": "Headers", "PreviouslyKnownMismatched": "MisMatched records in both files which is known", "FoundInOneFile": "Cells (rows) which are only found in 1010","FoundInTwoFile":"Cells (rows) which are only found in SEED"}
    sheetStats={"Matched":0,"SlightlyMatched":0,"MissMatched":0,"Headers":0,"PreviouslyKnownMismatched":0,"FoundInOneFile":0,"FoundInTwoFile":0}
    columnCountData={}
    def openFile(self,SheetTitle):
        self.wb = Workbook()
        self.workSheet = self.wb.active
        self.workSheet.title=SheetTitle

    def printIt(self,value,color):
        cell=self.workSheet.cell(row=self.rowCode, column=self.columnCode)
        #print value
        try:
            value = str(value).encode("ascii", errors="ignore")
        except Exception as e:
            print value
            value = str(value)
        try:

            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = value
        except Exception as e:
            print e
            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = "#SYSTEM ERROR"

        if (color != None):
            fill = PatternFill("solid", fgColor=color)
            cell.fill = fill
            if(self.columnCode not in self.columnCountData):
                self.columnCountData[self.columnCode]={}
            color_to_match = self.colType[color]
            if (color_to_match not in self.columnCountData[self.columnCode]):

                self.columnCountData[self.columnCode][color_to_match]=1
            else:
                self.columnCountData[self.columnCode][color_to_match] =self.columnCountData[self.columnCode][color_to_match] + 1

            self.columnCode = self.columnCode + 1

        if (color is None):
            print "r,c", self.rowCode, self.columnCode

    def printItLine(self, value=None,color=None):
        if(value!=None):
            cell = self.workSheet.cell(row=self.rowCode, column=self.columnCode)
            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = value
            if(color!=None):
                fill = PatternFill("solid", fgColor=color)
                cell.fill = fill
        self.rowCode = self.rowCode + 1
        self.columnCode = 1

    def createTableHeader(self, originalColumnsList,mappedDataColtoKey,mappedDataKeytoCol):
        for aColInOriList in originalColumnsList:
            self.printIt(aColInOriList,self.colourCode["Headers"])
            self.sheetStats["Headers"] = self.sheetStats["Headers"] + 1
        self.printItLine()

    def printLegendery(self):
        final_row_cunt = self.rowCode - 2
        print "final_row_cunt :- ",final_row_cunt
        self.printItLine()
        self.printItLine("Column wise data comparison %")
        for aoriColCountData in self.colourCode:
            for acolCountDataNum in self.columnCountData.keys():
                if aoriColCountData in self.columnCountData[acolCountDataNum]:
                    if "Headers" != aoriColCountData:
                        self.printIt(round(self.columnCountData[acolCountDataNum][aoriColCountData]*100.0/final_row_cunt,3),self.colourCode[aoriColCountData])
                    else:
                        self.printIt(0,self.colourCode[aoriColCountData])
                else:
                    self.printIt(0, self.colourCode[aoriColCountData])
            self.printItLine()
        self.printItLine()
        self.printIt("ColourCode",None)
        self.printIt("# Cells",None)
        self.printItLine("% Cells/Total Cells")
        colourKeyList=self.colourCode.keys()
        colourKeyList.sort()
        tot=0
        for aColourKey in colourKeyList:
            tot=tot+self.sheetStats[aColourKey]
        tot=tot-self.sheetStats["Headers"]
        print "Total row count ",tot
        for aColourKey in colourKeyList:
            self.printIt(self.colourCodeDesc[aColourKey], self.colourCode[aColourKey])
            self.printIt(self.sheetStats[aColourKey], self.colourCode[aColourKey])
            if(aColourKey!="Headers"):
                self.printItLine(round((self.sheetStats[aColourKey]*100.0 / tot),2),self.colourCode[aColourKey])
            else:
                self.printItLine()
        self.printItLine()
        self.printItLine(datacompareFooter)
        self.printItLine(datacompareDataFormat)
        self.printItLine()
        self.printItLine('KEYS:- '.join(colsToConsider))



    def saveIt(self,fileName):
        self.wb.save(fileName)



def readMappingFile(filename):
    originalColumnsList=[]
    mappedDataColtoKey = {}  # {'abc':2,'cd':2}
    mappedDataKeytoCol = {}  # {2:['abc','cd']}
    csvfile = open(filename)
    reader = csv.DictReader(csvfile)
    curRowNum=0
    headers = reader.fieldnames[:2]
    for row in reader:
        #print row
        temCols=[]
        for aHeader in headers:
            mappedDataColtoKey[row[aHeader]]=curRowNum
            temCols.append(row[aHeader])
            if(len(temCols)==1):
                originalColumnsList.append(row[aHeader])
        mappedDataKeytoCol[curRowNum]=temCols
        curRowNum=curRowNum+1
    return mappedDataColtoKey,mappedDataKeytoCol,originalColumnsList

def getRowsHashValue(row,columnMapperColtoKey,columnMapperKeytoCol):
    global colsToConsider
    MappedColsToConsider=[]  #create Mapping cols accroding to the considering Columns
    for aconsiderCol in colsToConsider:
        MappedColsToConsider.append(columnMapperColtoKey[aconsiderCol])#numbers of each column pair

    rowStringList = []
    for aMapColKey in MappedColsToConsider:
        MapColList=columnMapperKeytoCol[aMapColKey]
        for aMapCol in MapColList:
            try:
                try:
                    rowStringList.append(str(float(row[aMapCol])))
                except:
                    rowStringList.append(row[aMapCol])
            except:
                continue
    rowStringList.sort()
    #Basesd on the data of the considerable column rowwString and hash for that row may be created.
    rowString=""
    for data in rowStringList:
        rowString=rowString+data
    return rowString

def validate_date(date_text):
    try:
        dateis = datetime.datetime.strptime(date_text, '%Y-%m-%d')
        dateis = datetime.datetime.strftime(dateis, "%m/%d/%Y")
        return dateis
    except ValueError:
        if(date_text is not None and date_text.endswith(" ")):
            date_text_temp=date_text[:-1]
            try:
                dateis = datetime.datetime.strptime(date_text_temp, '%Y-%m-%d')

                dateis = str(int(datetime.datetime.strftime(dateis, "%m")))+"/"+datetime.datetime.strftime(dateis, "%d/%Y")
                #dateis = datetime.datetime.strftime(dateis,"%m/%d/%Y")
                return dateis
            except:
                pass
        return date_text

def getDataFromCSVFile(filename,mappedDataColtoKey,mappedDataKeytoCol):
    csvfile = open(filename)
    reader = csv.DictReader(csvfile)
    DataList = []
    DataHashList = []
    DataWithHashes={}
    for row in reader:
        DataList.append(row)
        hashVal=getRowsHashValue(row,mappedDataColtoKey,mappedDataKeytoCol)
        for k in row:
            try:
                row[k]=row[k].replace("$","").replace(",","").replace("%","")
                row[k]=validate_date(row[k])
            except:
                continue
        if hashVal in DataWithHashes:
            print DataWithHashes[hashVal]
            print row
            global secondaryPk
            if hashVal in DataWithHashes:
                print "step2 ", hashVal
                time.sleep(1)
                hashVal = hashVal + str(time.time()) + str(row)
                print hashVal
            #print "hash collided.-> ",hashVal
        DataHashList.append(hashVal)
        DataWithHashes[hashVal]=row
    return DataList,DataHashList,DataWithHashes

def is_number(s): ##ToDo refine code to work without try except
    try:
        num=float(s)
        return True,num
    except ValueError:
        return False,None

def similar(compString, compareeString,isString):
    if(isString):
        if compString==compareeString:
            return 1
        # elif float(compString) == float(compareeString):
        #     return 1
        return 0
    isCompStringNumber,CompStringdata=is_number(compString)
    iscompareeStringNumber,compareeStringdata = is_number(compareeString)

    if(isCompStringNumber and iscompareeStringNumber):#data are numbers
        if ((Decimal(CompStringdata) % 1 == 0) and (Decimal(compareeStringdata) % 1 == 0)):  ##If both value is integer then all consider as integers.
            CompStringdata = int(CompStringdata)  ##Conside data as integer
            compareeStringdata = int(compareeStringdata)
            if(CompStringdata==compareeStringdata):
                return 1.0
        elif ((Decimal(CompStringdata) % 1 == 0) or (Decimal(compareeStringdata) % 1 == 0)): ##If one value is integer then all consider as integers.
            CompStringdata = round(CompStringdata)  ##Conside data as integer
            compareeStringdata = round(compareeStringdata)
            #print CompStringdata,"===",compareeStringdata
        else:
            compareeStringdataDecimalPlaces=len(str(compareeStringdata).split(".")[1])
            CompStringdataDecimalPlaces=len(str(CompStringdata).split(".")[1])

            if(compareeStringdataDecimalPlaces>=2 and CompStringdataDecimalPlaces>=2):
                CompStringdata = round(CompStringdata, 2)
                compareeStringdata = round(compareeStringdata, 2)
                #if (CompStringdata == compareeStringdata):  # after rounding data matched
                    #return 1
            else:
                minDecimalPlaces=min(compareeStringdataDecimalPlaces,CompStringdataDecimalPlaces)
                CompStringdata = round(CompStringdata, minDecimalPlaces)
                compareeStringdata = round(compareeStringdata, minDecimalPlaces)

        if(CompStringdata==compareeStringdata):#after rounding data matched
            return 0.9
        else:
            return 0.4 #after rounding data didnt match
    else:
        if(compString=="" and (compareeString=="0" or compareeString=="0.0" or compareeString in ["0.00","0.000","0.000","0.0000"])):  #This data doesnt match we know it as exception
            return 1.1
        if(compString==compareeString):
            return 1
        if (compString.upper() == compareeString.upper()):
            return 1
        else:
            return 0


def getAndWriteMatchers(excellWriter,originalColumnsList,originalHashDataset,changedHashDataset,originalDataWithHashset,changedDataWithHashset,mappedDataColtoKey,mappedDataKeytoCol):
    global Stringcol,castToInt
    intersect=set(originalHashDataset).intersection(changedHashDataset)  ##Element common to both based on given column
    for element in intersect:
        oriDataRow= originalDataWithHashset[element]
        chngDataRow = changedDataWithHashset[element]
        for aColInOriList in originalColumnsList:
            isString = aColInOriList in Stringcol
            isCastToInt = aColInOriList in castToInt
            dataMatchList=mappedDataKeytoCol[mappedDataColtoKey[aColInOriList]]
            try:
                oriColData=oriDataRow[dataMatchList[0]]
                chngColData = chngDataRow[dataMatchList[1]]
            except:
                oriColData = oriDataRow[dataMatchList[1]]
                chngColData = chngDataRow[dataMatchList[0]]
            if isCastToInt:
                oriColData = str(int(float(oriColData)))
                chngColData = str(int(float(chngColData)))
            oriColDataOriginalData = oriColData
            chngColDataOriginalData = chngColData
            oriColData=oriColData.strip().replace(" ", "")
            chngColData = chngColData.strip().replace(" ", "")

            try:
                similarCompData=similar(oriColData, chngColData,isString)
            except:
                similarCompData = 0.4

            if(oriColData==chngColData or similarCompData==1):
                excellWriter.printIt(oriColDataOriginalData,excellWriter.colourCode["Matched"])
                excellWriter.sheetStats["Matched"]=excellWriter.sheetStats["Matched"]+1
            elif (similarCompData > 1.0):  ##May clear to go Detecting the previously known mismtches due to the technical difficulties
                excellWriter.printIt(oriColDataOriginalData + " != " + chngColDataOriginalData, excellWriter.colourCode["PreviouslyKnownMismatched"]) ##Clear for known exceptions
                excellWriter.sheetStats["PreviouslyKnownMismatched"] = excellWriter.sheetStats["PreviouslyKnownMismatched"] + 1
            elif(similarCompData>0.6): ##May clear to go
                excellWriter.printIt(oriColDataOriginalData+" != "+chngColDataOriginalData,excellWriter.colourCode["SlightlyMatched"])
                excellWriter.sheetStats["SlightlyMatched"] = excellWriter.sheetStats["SlightlyMatched"] + 1
            else:
                excellWriter.printIt(oriColDataOriginalData+" != "+chngColDataOriginalData,excellWriter.colourCode["MissMatched"])
                excellWriter.sheetStats["MissMatched"] = excellWriter.sheetStats["MissMatched"] + 1
        excellWriter.printItLine()
    return excellWriter

def getDiffAndWrite(compare,compareWith,compareHashedData,excellWriter,mappedDataColtoKey,mappedDataKeytoCol,originalColumnsList,colCode,file):
    firstDiff = set(compare).difference(compareWith)
    print "First diff ",len(firstDiff)
    for element in firstDiff:
        oriDataRow = compareHashedData[element]
        for aColInOriList in originalColumnsList:
            dataMatchList = mappedDataKeytoCol[mappedDataColtoKey[aColInOriList]]
            if file == "FoundInOneFile":
                excellWriter.printIt(oriDataRow[dataMatchList[0]],colCode)
                excellWriter.sheetStats["FoundInOneFile"] = excellWriter.sheetStats["FoundInOneFile"] + 1
            elif file == "FoundInTwoFile":
                excellWriter.printIt(oriDataRow[dataMatchList[1]],colCode)
                excellWriter.sheetStats["FoundInTwoFile"] = excellWriter.sheetStats["FoundInTwoFile"] + 1
            else:
                print "_____"

        excellWriter.printItLine()

    return excellWriter

def XmltoCSV(xmlFile):
    itm_nbr_postn=-1
    tree = ET.parse(xmlFile)
    root = tree.getroot()
    headerList= root.findall(".//th")

    rowList= root.findall(".//tr")

    outputfile=open('/home/nithilar/Desktop/xmltocsv.csv','wb')
    toWriteLine=""
    postnKey=0
    for aColName in headerList:
        if(aColName=="uid"):
            itm_nbr_postn=postnKey;
        toWriteLine=toWriteLine+","+aColName.get("name").encode('utf-8').replace("\n","").replace(",","")
        postnKey=postnKey+1
    outputfile.write(toWriteLine[1:]+"\n")


    for arow in rowList:
        toWriteLine=""
        i = 0
        dataCells= arow.findall(".//td")
        for acell in dataCells:
            if(acell.text==None):
                acell.text=""
            if(i==itm_nbr_postn):
                acell.text=str(int(str(acell.text.encode('utf-8').replace("\n","").replace(",",""))))
            toWriteLine=toWriteLine+","+str(acell.text.encode('utf-8').replace("\n","").replace(",",""))
            i=i+1
        outputfile.write(toWriteLine[1:] + "\n")
    outputfile.close()
    return outputfile.name


def columnMapFinder(colsFromJson,mappings,mainfileName, comparefileName): #Auto Column matcher
    bestMappedList = open(mapFile+".csv", "wb")
    bestMappedList.write(comparefileName+","+mainfileName+",MatchPrecentage\n")
    for aExtractedJsonCol in colsFromJson:
        matchedOne = ""
        matchedPrecentage = 0
        for amapping in mappings:
            matchval = similar(amapping, aExtractedJsonCol,True)
            if (matchedPrecentage < matchval):
                matchedPrecentage = matchval
                matchedOne = amapping + "," + aExtractedJsonCol
        if(matchedPrecentage==1.0):
            bestMappedList.write(matchedOne + "," + str(matchedPrecentage * 100) + "\n")
    firstDiff = set(colsFromJson).difference(mappings)
    print "First diff ",len(firstDiff)
    for element in firstDiff:
        bestMappedList.write( ","+element+"," + str(0) + "\n")
    firstDiff = set(mappings).difference(colsFromJson)
    print "First diff 2",len(firstDiff)
    for element in firstDiff:
        bestMappedList.write(element + ",," + str(0) + "\n")
    return bestMappedList.name

def checkSimilarityInteger(compString,compareeString):
    isCompStringNumber, CompStringdata = is_number(compString)
    iscompareeStringNumber, compareeStringdata = is_number(compareeString)
    if(isCompStringNumber and iscompareeStringNumber):
        if ((Decimal(CompStringdata) % 1 == 0) and (Decimal(compareeStringdata) % 1 == 0)):  ##If both value is integer then all consider as integers.
            CompStringdata = int(CompStringdata)  ##Conside data as integer
            compareeStringdata = int(compareeStringdata)
            if(CompStringdata==compareeStringdata):
                return 1.0
        elif (False and ((Decimal(CompStringdata) % 1 == 0) and Decimal(CompStringdata) != 0.0) or ((Decimal(compareeStringdata) % 1 == 0) and Decimal(CompStringdata) != 0.0)): ##If one value is integer then all consider as integers.
            CompStringdata = round(CompStringdata)  ##Conside data as integer
            compareeStringdata = round(compareeStringdata)
            print CompStringdata,"=!=",compareeStringdata
        else:
            compareeStringdataDecimalPlaces=len(str(compareeStringdata).split(".")[1])
            CompStringdataDecimalPlaces=len(str(CompStringdata).split(".")[1])

            if(compareeStringdataDecimalPlaces>=2 and CompStringdataDecimalPlaces>=2):
                CompStringdata = round(CompStringdata, 2)
                compareeStringdata = round(compareeStringdata, 2)
                #if (CompStringdata == compareeStringdata):  # after rounding data matched
                    #return 1
            else:
                minDecimalPlaces=min(compareeStringdataDecimalPlaces,CompStringdataDecimalPlaces)
                CompStringdata = round(CompStringdata, minDecimalPlaces)
                compareeStringdata = round(compareeStringdata, minDecimalPlaces)

        if(CompStringdata==compareeStringdata):#after rounding data matched
            return 0.9
        else:
            return 0.4 #after rounding data didnt match


def checkSimilarityString(compString,compareeString):
    if(compString==compareeString):
        return 1
    if (compString == "0" and compareeString == ""):  # This data doesnt match we know it as exception
        return 1.1
    return SequenceMatcher(None, compString, compareeString).ratio()  ##Finally string matching


def ColumnMapGenerator(mainfileName, comparefileName):  # Auto Column matcher
    data_file = open(mainfileName)
    colsFromJson = data_file.readline().strip().split(",")
    data_file.close()
    configFile = open(comparefileName)
    configLines = configFile.readline()
    mappings = configLines.strip().split(",")
    configFile.close()
    return columnMapFinder(colsFromJson, mappings,mainfileName, comparefileName)


def dataCompare(xmlFile,csvFile,file_nm):
    xmlTocsvFile=xmlFile
    mappedDataColtoKey,mappedDataKeytoCol,originalColumnsList=readMappingFile(mapFile+".csv")

    print "Loading started for {}".format(csvFile)
    originalDataset, originalHashDataset, originalDataWithHashset = getDataFromCSVFile(csvFile,mappedDataColtoKey,mappedDataKeytoCol)
    print "Loading started for {}".format(xmlFile)
    changedDataset, changedHashDataset, changedDataWithHashset = getDataFromCSVFile(xmlTocsvFile,mappedDataColtoKey,mappedDataKeytoCol)
    print "Data loading completed"
    ####variable description########
    # columnMapper - Hold the mapping as declared in the ColumnMap.csv file
    # originalColumnsList - It will  hold the order which data must be represented
    # originalDataset | changedDataset - raw data of each row
    # originalHashDataset | changedHashDataset- list of hash value for the data set Based on gove columns
    # originalDataWithHashset | changedDataWithHashset- hash value and dataSet mapping


    ###Calculate and data writing
    excellWriter = WriteToExcel()
    excellWriter.openFile("Data")
    excellWriter.colourCodeDesc["FoundInOneFile"] ="Cells (rows) which are only found in "+xmlFile_nm
    excellWriter.colourCodeDesc["FoundInTwoFile"] ="Cells (rows) which are only found in "+csvFile_nm
    excellWriter.createTableHeader(originalColumnsList,mappedDataColtoKey,mappedDataKeytoCol)

    getAndWriteMatchers(excellWriter, originalColumnsList, originalHashDataset, changedHashDataset,
                        originalDataWithHashset, changedDataWithHashset,mappedDataColtoKey,mappedDataKeytoCol)
    getDiffAndWrite(originalHashDataset, changedHashDataset, originalDataWithHashset, excellWriter, mappedDataColtoKey, mappedDataKeytoCol,
                    originalColumnsList, excellWriter.colourCode["FoundInOneFile"],"FoundInOneFile")
    getDiffAndWrite(changedHashDataset, originalHashDataset, changedDataWithHashset, excellWriter, mappedDataColtoKey, mappedDataKeytoCol,
                    originalColumnsList, excellWriter.colourCode["FoundInTwoFile"],"FoundInTwoFile")
    excellWriter.printLegendery()
    excellWriter.printItLine()
    excellWriter.saveIt(file_nm)
    #excellWriter.saveIt("/home/nithilar/PycharmProjects/DataCompare/DataCompare02/hdata0003/msr_1010.xls")



'''
xmlFile="/home/nithilar/PythonBackup/DataValidate/dataSet/hh08/1010.csv"
csvFile="/home/nithilar/PythonBackup/DataValidate/dataSet/hh08/msr.csv"
mapFile="/home/nithilar/PythonBackup/DataValidate/dataSet/hh08/map.csv"
file_nm="/home/nithilar/PythonBackup/DataValidate/dataSet/hh08/out.xlsx"
'''

if (len(sys.argv)==6 and (sys.argv[1]=="m")):
    ismap = True
    xmlFile = sys.argv[2]
    csvFile = sys.argv[3]
    mapFile = sys.argv[4]
    file_nm = sys.argv[5]
else:
    ismap = False
    xmlFile = sys.argv[1]
    csvFile = sys.argv[2]
    mapFile = sys.argv[3]
    file_nm = sys.argv[4]
    if(len(sys.argv) > 5):
    	pkColStr = sys.argv[5]
    	colsToConsider = pkColStr.split(",")
    print "python DataValidate.py "+xmlFile+" "+csvFile+" "+mapFile+ " "+file_nm

if len(xmlFile.split("/"))==1:
    xmlFile_nm= xmlFile
else:
    xmlFile_nm=xmlFile.split("/")[len(xmlFile.split("/"))-1]

if len(csvFile.split("/"))==1:
    csvFile_nm= csvFile
else:
    csvFile_nm=csvFile.split("/")[len(csvFile.split("/"))-1]

datacompareFooter = "Data Comparison:- "+xmlFile_nm+" ~ "+csvFile_nm
datacompareDataFormat = "Data Format for mismatch cells <"+xmlFile_nm+"> != <"+csvFile_nm+">"


if(ismap):
 #xmlTocsvFile = XmltoCSV(xmlFile4
 ColumnMapGenerator(csvFile, xmlFile)
 print "---"
else:
 dataCompare(csvFile, xmlFile,file_nm)

print("======================================================================")
print("============== PLEASE CHECK THE FILE NAMES ARE SHAREABLE =============")
print(xmlFile_nm)
print(csvFile_nm)
print("======================================================================")