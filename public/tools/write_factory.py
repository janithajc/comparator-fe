from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging
import csv
import os
import time

def copy_csv_file_with_header_manipulation(input_file):
    print "HEADER CLENSING STARTED FOR {0}".format(input_file)
    output_file = input_file+"_header_temp.csv"
    with open(input_file, 'rb') as inFile, open(output_file, 'wb') as outfile:
        header_row = inFile.readline().replace("%","_pct")
        outfile.write(header_row)
        for i in inFile:
            outfile.write(i)
    os.remove(input_file)
    os.rename(output_file, input_file)
    print "HEADER CLENSING COMPLETED {0}".format(input_file)


class WriteToExcel():
    workSheet=None
    columnCode=1
    rowCode=1
    wb=None
    def __init__(self,colourCode,colourCodeDesc):
        self.colourCode=colourCode
        self.colourCodeDesc=colourCodeDesc
    columnCountData={}
    def openFile(self,SheetTitle):
        self.wb = Workbook()
        self.workSheet = self.wb.active
        self.workSheet.title=SheetTitle

    def printIt(self,value,color):
        cell=self.workSheet.cell(row=self.rowCode, column=self.columnCode)
        #print value
        try:
            value = str(value).encode("ascii", errors="ignore")
        except Exception as e:
            print value
            value = str(value)
        try:

            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = value
        except Exception as e:
            print e
            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = "#SYSTEM ERROR"

        if (color != None):
            fill = PatternFill("solid", fgColor=color)
            cell.fill = fill
            self.columnCode = self.columnCode + 1

        if (color is None):
            print "r,c", self.rowCode, self.columnCode

    def printItLine(self, value=None,color=None):
        if(value!=None):
            cell = self.workSheet.cell(row=self.rowCode, column=self.columnCode)
            self.workSheet.cell(row=self.rowCode, column=self.columnCode).value = value
            if(color!=None):
                fill = PatternFill("solid", fgColor=color)
                cell.fill = fill
        self.rowCode = self.rowCode + 1
        self.columnCode = 1

    def createTableHeader(self, originalColumnsList):
        for aColInOriList in originalColumnsList:
            self.printIt(aColInOriList,self.colourCode["Headers"])
        self.printItLine()

    def printLegendery(self,legend_recorder,column_list):
        final_row_cunt = self.rowCode - 2
        summarized_data={}
        print final_row_cunt
        self.printItLine()
        self.printItLine("Column wise summary")
        col_keys =self.colourCode.keys()
        col_keys.remove("Headers")
        for j in col_keys:
            for i in column_list:
                if self.colourCodeDesc[j]=="Headers":
                    continue
                self.printIt(round(legend_recorder[i][j]*100.0/final_row_cunt,2),self.colourCode[j])
                if j in summarized_data:
                    summarized_data[j]=summarized_data[j]+legend_recorder[i][j]
                else:
                    summarized_data[j]=legend_recorder[i][j]
            self.printItLine()
        self.printItLine()
        self.printItLine("Summarized data in cell wise")
        self.printIt("Colour desc","5D6D7E")
        self.printIt("Cell count","5D6D7E")
        self.printIt("Cell %","5D6D7E")
        self.printItLine()
        for j in col_keys:
            if self.colourCodeDesc[j]=="Headers":
                continue
            self.printIt(self.colourCodeDesc[j],self.colourCode[j])
            self.printIt(summarized_data[j],self.colourCode[j])
            self.printIt(round(summarized_data[j]*100.0/(final_row_cunt*len(column_list)),2),self.colourCode[j])
            self.printItLine()
            
    def saveIt(self,fileName):
        self.wb.save(fileName)
